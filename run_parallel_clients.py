"""
PARALLEL Batch Runner — Run 100 clients in parallel (5 at once).
Each client runs in its own isolated config — no conflict.

Usage:
  python run_parallel_clients.py
  python run_parallel_clients.py --workers 10   # 10 parallel
  python run_parallel_clients.py --smoke        # run only E2E tests (faster)

Input: data/clients.json (list of clients)
Output: reports/Client_Regression_Report_TIMESTAMP.xlsx + per-client allure reports
"""

import json
import os
import re
import sys
import subprocess
import argparse
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl"], check=True)
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


PROJECT_DIR = os.path.dirname(os.path.abspath(__file__))
CLIENTS_FILE = os.path.join(PROJECT_DIR, "data", "clients.json")
REPORTS_DIR = os.path.join(PROJECT_DIR, "reports")
TIMESTAMP = datetime.now().strftime("%Y%m%d_%H%M%S")


# ── Canonical columns for the detail sheets ─────────────────────────
WALLET_COLUMNS = [
    "PhonePe", "AmazonPay", "MobiKwik", "Airtel Money",
    "FreeCharge", "Jio", "OLA Money",
]

NETBANKING_COLUMNS = [
    "HDFC Bank", "ICICI Bank", "State Bank of India", "Axis Bank",
    "Kotak Mahindra Bank", "IDFC FIRST Bank Limited",
    "Punjab National Bank [Retail]", "Bank of Baroda Net Banking Corporate",
    "Bank of Baroda Net Banking Retail", "Indusind Bank", "Fedral Bank",
    "Karur Vysya Bank", "Indian Overseas Bank", "RBL Bank",
    "Standard Chartered Bank", "Deutsche Bank", "Jammu and Kashmir Bank",
    "Saraswat Bank", "Dhanlaxmi Bank", "City Union Bank", "Karnataka Bank",
    "Lakshmi Vilas Bank NetBanking", "Equitas Bank", "Catholic Syrian Bank",
    "Janata Sahakari Bank LTD Pune", "Punjab & Sind Bank",
    "Punjab & Maharashtra Co-op Bank Ltd", "UCO Bank",
    "Tamilnad Mercantile Bank", "DBS Bank Ltd", "Royal Bank Of Scotland",
]

OFFLINE_COLUMNS = [
    "Cash → ICICI Bank", "Cash → Airtel Bank", "Cash → FINO Bank",
    "Cash → Bank of India Retail",
    "RTGS → IDFC First Bank RTGS",
    "IMPS → SabPaisa (SA)",
]

CARD_COLUMNS = [
    "Debit Card", "Credit Card",
]

# Test-name → card-column map (case-insensitive substring on test fullName)
CARD_TEST_MAP = {
    "debit_card": "Debit Card",
    "credit_card": "Credit Card",
    "debit card": "Debit Card",
    "credit card": "Credit Card",
}


def load_clients():
    with open(CLIENTS_FILE, "r") as f:
        return json.load(f)


def run_client(client, smoke_only=False):
    """Run tests for single client in isolated env vars."""
    mid = client["merchant_id"]
    env = client["environment"]
    allure_dir = os.path.join(REPORTS_DIR, f"allure-results-{mid}")

    # Pass config via env vars — no shared file conflict
    env_vars = os.environ.copy()
    env_vars["MERCHANT_ID_OVERRIDE"] = mid
    env_vars["ENV"] = env
    env_vars["HEADLESS"] = "true"
    env_vars["SLOW_MO"] = "0"

    cmd = [
        sys.executable, "-m", "pytest",
        "tests/test_regression_suite.py",
        f"--alluredir={allure_dir}",
        "--clean-alluredir",
        "-q",
    ]

    if smoke_only:
        cmd.extend(["-k", "R9 or R10"])  # only E2E + full sequential

    start = datetime.now()
    print(f"[START] {mid} ({env})")

    try:
        result = subprocess.run(
            cmd, cwd=PROJECT_DIR, env=env_vars,
            capture_output=True, text=True, timeout=600,
        )
        duration = (datetime.now() - start).total_seconds()
        print(f"[DONE ] {mid}: exit={result.returncode} in {duration:.0f}s")
        return client, allure_dir, result.returncode, duration
    except subprocess.TimeoutExpired:
        print(f"[TIMEOUT] {mid}")
        return client, allure_dir, -1, 600


def _extract_items_from_attachment(text):
    """Parse per-sub-item status lines from an attachment body.
    Supports two formats produced by the test suite:
      - '[OK]   Name ...'    / '[PASS]  Name ...'  / '[FAIL] Name: ...'
      - 'Name: PASS'         / 'Name: FAIL'
    Returns dict: {canonical_item_name: 'PASS'|'FAIL'}.
    """
    out = {}
    for raw in text.splitlines():
        line = raw.strip()
        if not line:
            continue

        # Format A: [TAG] item...
        m = re.match(r"^\[(OK|PASS|FAIL)\]\s+(.+?)(?::|$)", line)
        if m:
            tag = m.group(1)
            item = m.group(2).strip()
            # Strip leading step numbers like "1. " or "12. "
            item = re.sub(r"^\d+\.?\s*", "", item).strip()
            out[item] = "PASS" if tag in ("OK", "PASS") else "FAIL"
            continue

        # Format B: "Name: PASS" / "Name: FAIL"
        m = re.match(r"^(.+?):\s*(PASS|FAIL)\s*$", line)
        if m:
            item = m.group(1).strip()
            out[item] = m.group(2).upper()
    return out


def _match_canonical(item, canonical_list):
    """Find which canonical column an item belongs to.
    Case-insensitive exact first, then substring both directions.
    Returns the canonical column name or None."""
    low = item.lower()
    for col in canonical_list:
        if col.lower() == low:
            return col
    for col in canonical_list:
        if col.lower() in low or low in col.lower():
            return col
    return None


def parse_results(allure_dir):
    results = {
        "total": 0, "passed": 0, "failed": 0, "broken": 0,
        "UPI": "N/A", "Cards": "N/A", "Netbanking": "N/A",
        "Wallets": "N/A", "Offline": "N/A",
        "Config": "N/A", "Customer": "N/A",
        # per-sub-item detail dicts, keyed by canonical column name
        "_wallet_detail": {},
        "_netbanking_detail": {},
        "_offline_detail": {},
        "_card_detail": {},
    }
    mode_map = {
        "upi": "UPI", "qr": "UPI",
        "card": "Cards", "cvv": "Cards",
        "netbanking": "Netbanking", "bank": "Netbanking", "equitas": "Netbanking",
        "wallet": "Wallets", "phonepe": "Wallets",
        "offline": "Offline", "cash": "Offline",
        "config": "Config", "merchant": "Config", "fetch": "Config",
        "customer": "Customer",
    }
    if not os.path.exists(allure_dir):
        return results
    for fname in os.listdir(allure_dir):
        if not fname.endswith("-result.json"):
            continue
        try:
            with open(os.path.join(allure_dir, fname), "r", encoding="utf-8") as f:
                data = json.load(f)
            status = data.get("status", "unknown")
            name = (data.get("name", "") + " " + data.get("fullName", "")).lower()
            results["total"] += 1
            if status == "passed":
                results["passed"] += 1
            elif status == "failed":
                results["failed"] += 1
            elif status == "broken":
                results["broken"] += 1
            matched_mode = None
            for keyword, mode in mode_map.items():
                if keyword in name:
                    matched_mode = mode
                    if results[mode] == "N/A":
                        results[mode] = status.upper()
                    elif results[mode] == "PASSED" and status != "passed":
                        results[mode] = status.upper()
                    break

            # Map specific test names (like *_debit_card / *_credit_card)
            # into the Cards Detail columns, test-status directly.
            for keyword, card_col in CARD_TEST_MAP.items():
                if keyword in name:
                    card_status = "PASS" if status == "passed" else "FAIL"
                    _merge_detail(results["_card_detail"], card_col, card_status)
                    break

            # ── Pull per-sub-item detail from .txt attachments ─────
            for att in data.get("attachments", []):
                source = att.get("source", "")
                if not source.endswith(".txt"):
                    continue
                att_path = os.path.join(allure_dir, source)
                if not os.path.exists(att_path):
                    continue
                try:
                    with open(att_path, "r", encoding="utf-8", errors="ignore") as af:
                        body = af.read()
                except Exception:
                    continue
                items = _extract_items_from_attachment(body)
                if not items:
                    continue

                # Map each parsed item to a canonical detail bucket
                att_name = att.get("name", "").lower()
                for item, st in items.items():
                    # Try Wallets
                    if matched_mode == "Wallets" or "wallet" in att_name:
                        col = _match_canonical(item, WALLET_COLUMNS)
                        if col:
                            _merge_detail(results["_wallet_detail"], col, st)
                            continue
                    # Try Netbanking
                    if matched_mode == "Netbanking" or "bank" in att_name:
                        col = _match_canonical(item, NETBANKING_COLUMNS)
                        if col:
                            _merge_detail(results["_netbanking_detail"], col, st)
                            continue
                    # Try Offline
                    if matched_mode == "Offline" or "offline" in att_name:
                        col = _match_canonical(item, OFFLINE_COLUMNS)
                        if col:
                            _merge_detail(results["_offline_detail"], col, st)
                            continue
        except Exception:
            pass
    return results


def _merge_detail(detail_dict, col, status):
    """Merge a per-item status into the detail dict. FAIL wins over PASS
    (so a bank that failed in one test stays FAIL even if passed elsewhere)."""
    existing = detail_dict.get(col)
    if existing is None:
        detail_dict[col] = status
    elif existing == "PASS" and status == "FAIL":
        detail_dict[col] = "FAIL"


def generate_allure_html_per_client(clients):
    """For each client, convert reports/allure-results-<MID>/ into a
    standalone HTML report at reports/allure-report-<MID>/.
    Returns list of (merchant_id, html_path, status)."""
    out = []
    for client in clients:
        mid = client["merchant_id"]
        results_dir = os.path.join(REPORTS_DIR, f"allure-results-{mid}")
        report_dir = os.path.join(REPORTS_DIR, f"allure-report-{mid}")
        if not os.path.isdir(results_dir) or not os.listdir(results_dir):
            out.append((mid, None, "no-results"))
            continue
        try:
            # shell=True lets Windows resolve allure.bat transparently
            subprocess.run(
                ["allure", "generate", results_dir, "-o", report_dir, "--clean"],
                check=True, capture_output=True, text=True, shell=True,
            )
            out.append((mid, report_dir, "ok"))
        except subprocess.CalledProcessError as e:
            out.append((mid, None, f"error: {e.stderr[:80] if e.stderr else e}"))
        except FileNotFoundError:
            out.append((mid, None, "allure CLI not found on PATH"))
    return out


def _col_letter(n):
    """1 → A, 27 → AA, etc. Supports more than 26 columns."""
    s = ""
    while n > 0:
        n, rem = divmod(n - 1, 26)
        s = chr(65 + rem) + s
    return s


def _make_detail_sheet(wb, sheet_name, columns, all_results, detail_key, pass_fill, fail_fill, na_fill, header_font, header_fill, border, center):
    """Build a detail sheet: rows = clients, columns = each canonical sub-item."""
    ws = wb.create_sheet(sheet_name)
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1, end_column=3 + len(columns))
    ws.cell(row=1, column=1,
            value=f"{sheet_name} — per-item status (PASS / FAIL / N/A)").font = \
        Font(bold=True, size=12, color="1565C0")
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")

    headers = ["S.No", "Client Code", "Env"] + columns
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = border

    for idx, (client, res, _duration) in enumerate(all_results, 1):
        row = idx + 3
        detail = res.get(detail_key, {})
        ws.cell(row=row, column=1, value=idx).alignment = center
        ws.cell(row=row, column=2, value=client["merchant_id"]).alignment = center
        ws.cell(row=row, column=3, value=client["environment"]).alignment = center
        for k, col_name in enumerate(columns):
            status = detail.get(col_name, "N/A")
            c = ws.cell(row=row, column=4 + k, value=status)
            c.alignment = center
            c.border = border
            if status == "PASS":
                c.fill = pass_fill
            elif status == "FAIL":
                c.fill = fail_fill
            else:
                c.fill = na_fill

    # Widths
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 12
    for k in range(len(columns)):
        ws.column_dimensions[_col_letter(4 + k)].width = 22


def generate_excel(all_results, total_duration):
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill("solid", start_color="1565C0")
    pass_fill = PatternFill("solid", start_color="C6EFCE")
    fail_fill = PatternFill("solid", start_color="FFC7CE")
    na_fill = PatternFill("solid", start_color="F5F5F5")
    border = Border(*[Side(style="thin")] * 4)
    center = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A1:N1")
    ws["A1"] = f"SabPaisa Regression - {len(all_results)} Clients - {datetime.now().strftime('%d-%b-%Y %H:%M')} - Duration: {total_duration:.0f}s"
    ws["A1"].font = Font(bold=True, size=13, color="1565C0")
    ws["A1"].alignment = Alignment(horizontal="center")

    headers = ["S.No", "Client Code", "Env", "Total", "Passed", "Failed", "Broken",
               "Config", "UPI", "Cards", "Netbanking", "Wallets", "Offline", "Status"]

    for col, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = border

    for idx, (client, res, duration) in enumerate(all_results, 1):
        row = idx + 3
        is_pass = res["failed"] == 0 and res["broken"] == 0 and res["total"] > 0
        overall = "PASS" if is_pass else ("FAIL" if res["total"] > 0 else "NOT RUN")

        values = [idx, client["merchant_id"], client["environment"],
                  res["total"], res["passed"], res["failed"], res["broken"],
                  res["Config"], res["UPI"], res["Cards"],
                  res["Netbanking"], res["Wallets"], res["Offline"], overall]

        for col, v in enumerate(values, 1):
            c = ws.cell(row=row, column=col, value=v)
            c.alignment = center
            c.border = border
            if isinstance(v, str):
                if "PASS" in v:
                    c.fill = pass_fill
                elif "FAIL" in v or "BROKEN" in v:
                    c.fill = fail_fill
                elif v == "N/A" or v == "NOT RUN":
                    c.fill = na_fill

    # Summary
    summary_row = len(all_results) + 5
    total = len(all_results)
    passed = sum(1 for _, r, _ in all_results if r["failed"] == 0 and r["broken"] == 0 and r["total"] > 0)

    ws.cell(row=summary_row, column=1, value="SUMMARY").font = Font(bold=True, size=12)
    ws.cell(row=summary_row + 1, column=1, value="Total Clients:")
    ws.cell(row=summary_row + 1, column=2, value=total)
    ws.cell(row=summary_row + 2, column=1, value="Passed:")
    ws.cell(row=summary_row + 2, column=2, value=passed).fill = pass_fill
    ws.cell(row=summary_row + 3, column=1, value="Failed:")
    ws.cell(row=summary_row + 3, column=2, value=total - passed).fill = fail_fill if total - passed > 0 else na_fill
    ws.cell(row=summary_row + 4, column=1, value="Total Duration:")
    ws.cell(row=summary_row + 4, column=2, value=f"{total_duration:.0f} seconds ({total_duration/60:.1f} min)")

    widths = [6, 15, 12, 7, 8, 8, 8, 10, 8, 8, 12, 10, 10, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    # ── Detail sheets: Cards / Netbanking / Wallets / Offline ──────
    _make_detail_sheet(wb, "Cards Detail", CARD_COLUMNS, all_results,
                       "_card_detail", pass_fill, fail_fill, na_fill,
                       header_font, header_fill, border, center)
    _make_detail_sheet(wb, "Netbanking Detail", NETBANKING_COLUMNS, all_results,
                       "_netbanking_detail", pass_fill, fail_fill, na_fill,
                       header_font, header_fill, border, center)
    _make_detail_sheet(wb, "Wallets Detail", WALLET_COLUMNS, all_results,
                       "_wallet_detail", pass_fill, fail_fill, na_fill,
                       header_font, header_fill, border, center)
    _make_detail_sheet(wb, "Offline Detail", OFFLINE_COLUMNS, all_results,
                       "_offline_detail", pass_fill, fail_fill, na_fill,
                       header_font, header_fill, border, center)

    path = os.path.join(REPORTS_DIR, f"Client_Regression_Report_{TIMESTAMP}.xlsx")
    os.makedirs(REPORTS_DIR, exist_ok=True)
    wb.save(path)
    return path


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--workers", type=int, default=5, help="Parallel workers (default: 5)")
    parser.add_argument("--smoke", action="store_true", help="Run only E2E/smoke (R9+R10) - faster")
    parser.add_argument("--clients", type=str, default=None,
                        help="Comma-separated merchant IDs to run (e.g. CHIN36,SUBI79). "
                             "If omitted, runs every client in data/clients.json.")
    args = parser.parse_args()

    clients = load_clients()

    # Optional filter by --clients CHIN36,SUBI79
    if args.clients:
        wanted = {c.strip().upper() for c in args.clients.split(",") if c.strip()}
        clients = [c for c in clients if c["merchant_id"].upper() in wanted]
        missing = wanted - {c["merchant_id"].upper() for c in clients}
        if missing:
            print(f"WARNING: These merchant IDs not found in clients.json: {', '.join(sorted(missing))}")
        if not clients:
            print("ERROR: No matching clients. Check data/clients.json.")
            sys.exit(1)

    print(f"\n{'='*60}")
    print(f"  PARALLEL CLIENT RUNNER")
    print(f"{'='*60}")
    print(f"  Clients:  {len(clients)} ({', '.join(c['merchant_id'] for c in clients)})")
    print(f"  Workers:  {args.workers} (parallel)")
    print(f"  Mode:     {'SMOKE (R9+R10 only)' if args.smoke else 'FULL REGRESSION'}")
    print(f"{'='*60}\n")

    start = datetime.now()
    all_results = []

    with ThreadPoolExecutor(max_workers=args.workers) as pool:
        futures = {pool.submit(run_client, c, args.smoke): c for c in clients}
        for fut in as_completed(futures):
            client, allure_dir, rc, duration = fut.result()
            results = parse_results(allure_dir)
            all_results.append((client, results, duration))

    # Sort by merchant_id for consistent order
    all_results.sort(key=lambda x: x[0]["merchant_id"])

    total_duration = (datetime.now() - start).total_seconds()
    excel_path = generate_excel(all_results, total_duration)

    # Generate static Allure HTML reports per client
    print(f"\n  Generating Allure HTML reports per client...")
    allure_reports = generate_allure_html_per_client(clients)

    # Final summary
    total = len(all_results)
    passed = sum(1 for _, r, _ in all_results if r["failed"] == 0 and r["broken"] == 0 and r["total"] > 0)

    print(f"\n{'='*60}")
    print(f"  BATCH COMPLETE")
    print(f"{'='*60}")
    print(f"  Total Clients: {total}")
    print(f"  Passed:        {passed}")
    print(f"  Failed:        {total - passed}")
    print(f"  Duration:      {total_duration:.0f} seconds ({total_duration/60:.1f} min)")
    print(f"  Excel Report:  {excel_path}")
    print(f"\n  Allure HTML reports (one per client):")
    for mid, path, status in allure_reports:
        if path:
            print(f"    {mid:<12} -> {path}")
            print(f"    {' '*12}    open: allure open \"{path}\"")
        else:
            print(f"    {mid:<12} -> [{status}]")
    print(f"{'='*60}\n")


if __name__ == "__main__":
    main()
