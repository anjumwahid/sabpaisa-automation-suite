"""
Convert Excel sheet of client codes to clients.json

Usage:
  python excel_to_clients.py manager_clients.xlsx
  python excel_to_clients.py manager_clients.xlsx --env Staging
  python excel_to_clients.py manager_clients.xlsx --env Production

Expected Excel format (column names flexible):
  | Client Code | Environment |
  | UTTA99      | Staging     |
  | SQUA102     | Production  |
  | ...         | ...         |

If only 1 column (just client codes), pass --env to set environment for all.
"""

import sys
import json
import os
import argparse

try:
    from openpyxl import load_workbook
except ImportError:
    import subprocess
    subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl"], check=True)
    from openpyxl import load_workbook


PROJECT_DIR = os.path.dirname(os.path.abspath(__file__))
CLIENTS_JSON = os.path.join(PROJECT_DIR, "data", "clients.json")


def find_column(headers, keywords):
    """Find column index by keyword match."""
    for i, h in enumerate(headers):
        if h and any(k.lower() in str(h).lower() for k in keywords):
            return i
    return None


def convert(excel_path, default_env="Staging"):
    wb = load_workbook(excel_path, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        print("ERROR: Excel is empty")
        return

    # Find header row (first row with any value)
    header_idx = 0
    for i, row in enumerate(rows):
        if any(cell for cell in row):
            header_idx = i
            break

    headers = rows[header_idx]
    data_rows = rows[header_idx + 1:]

    # Find columns
    code_col = find_column(headers, ["client", "merchant", "code", "id"])
    env_col = find_column(headers, ["env", "environment", "stage"])

    if code_col is None:
        # Assume first non-empty column is client code
        code_col = 0
        print(f"No 'Client Code' header found - using column A")

    print(f"\nReading Excel: {excel_path}")
    print(f"  Client Code column: {chr(65 + code_col)} ({headers[code_col]})")
    if env_col is not None:
        print(f"  Environment column: {chr(65 + env_col)} ({headers[env_col]})")
    else:
        print(f"  Environment column: Not found - using default: {default_env}")

    clients = []
    for row in data_rows:
        if not row or not row[code_col]:
            continue
        code = str(row[code_col]).strip()
        if not code or code.lower() in ["none", "null", ""]:
            continue

        env = default_env
        if env_col is not None and row[env_col]:
            env = str(row[env_col]).strip()

        clients.append({"merchant_id": code, "environment": env})

    # Save to clients.json
    with open(CLIENTS_JSON, "w") as f:
        json.dump(clients, f, indent=4)

    print(f"\nConverted {len(clients)} clients:")
    for i, c in enumerate(clients[:5], 1):
        print(f"  {i}. {c['merchant_id']} ({c['environment']})")
    if len(clients) > 5:
        print(f"  ... and {len(clients) - 5} more")

    print(f"\nSaved to: {CLIENTS_JSON}")
    print(f"\nNow run:")
    print(f"  python run_parallel_clients.py --workers 5 --smoke")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("excel_file", help="Path to Excel file from manager")
    parser.add_argument("--env", default="Staging", help="Default environment if not in Excel (default: Staging)")
    args = parser.parse_args()

    if not os.path.exists(args.excel_file):
        print(f"ERROR: File not found: {args.excel_file}")
        sys.exit(1)

    convert(args.excel_file, args.env)


if __name__ == "__main__":
    main()
